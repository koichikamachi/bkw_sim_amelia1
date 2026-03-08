# ==============================
#  bkw_sim_amelia1/ui/app.py
#  仕様書v4訂正済準拠版
#  入力禁則・依存制御・免責同意・シナリオCSV・結果Excelダウンロード
#  サイドバー expander方式
# ==============================

import os
import sys

current_dir  = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import streamlit as st
import pandas as pd
import numpy as np
import datetime
import traceback
from io import BytesIO
from typing import List

from config.params import (
    SimulationParams,
    LoanParams,
    ExitParams,
    AdditionalInvestmentParams,
)
from core.simulation.simulation import Simulation
from core.finance.fs_builder import FinancialStatementBuilder


# ============================================================
# CF 表示ラベル変換マップ（fs_builder内部ラベル → UI/Excel表示ラベル）
# 7) 税込ラベルへの変換
# ============================================================
CF_LABEL_MAP = {
    "家賃収入（税抜）":           "家賃収入（税込）",
    "管理費・修繕費・保険料":     "管理費・修繕費・保険料（税込）",
    "売却費用":                   "売却費用（税込）",
    "土地購入":                   "土地購入（非課税）",
    "建物購入":                   "建物購入（税込）",
    "追加設備購入":               "追加設備購入（税込）",
    "固定資産売却収入":           "固定資産売却収入（税込）",
}

# 6) PL 表示ラベル変換マップ
PL_LABEL_MAP = {
    "販売費一般管理費": "管理費",
    "修繕費":           "修繕費",
    "保険料":           "保険料",
    "その他販管費":     "その他販管費",
}


# ============================================================
# CSS
# 3) .bkw-label の color を #2c3e50（探偵レポートと同色）に統一
# ============================================================
def inject_global_css():
    st.markdown(
        """
        <style>
        .bkw-card {
            background-color: #f4f5f7;
            border-left: 4px solid #2c3e50;
            padding: 10px 14px;
            margin-bottom: 8px;
            border-radius: 8px;
        }
        .bkw-label {
            font-size: 0.95rem;
            font-weight: 700;
            color: #1a2744;
            margin-bottom: 2px;
        }
        .bkw-value {
            font-size: 1.05rem;
            font-weight: 800;
            color: #111;
            text-align: right;
            font-variant-numeric: tabular-nums;
        }
        .bkw-section-title {
            font-size: 1.25rem;
            font-weight: 800;
            margin-top: 26px;
            margin-bottom: 14px;
            color: #2c3e50;
        }
        div.stButton > button {
            font-size: 1.1rem !important;
            font-weight: 800 !important;
            padding: 0.6em 1.1em !important;
        }
        .stTabs [data-baseweb="tab"] { font-size: 0.95rem; }
        </style>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# 数値フォーマット
# ============================================================
def _fmt(val):
    if pd.isna(val) or (isinstance(val, float) and np.isnan(val)):
        return ""
    if isinstance(val, (int, float, np.integer, np.floating)):
        try:
            return f"{int(round(val)):,}"
        except Exception:
            return str(val)
    return str(val)


def _apply_label_map(df: pd.DataFrame, lmap: dict) -> pd.DataFrame:
    df = df.copy()
    df.index = [lmap.get(i, i) for i in df.index]
    return df


# ============================================================
# 表示用 DataFrame 生成
# 5) 数値列を右寄せスタイル付きで返す
# ============================================================
def create_display_dataframes(fs_data: dict) -> dict:
    out = {}
    for key in ["pl", "bs", "cf"]:
        if key not in fs_data:
            continue
        df = fs_data[key].copy()
        if key == "cf":
            df = _apply_label_map(df, CF_LABEL_MAP)   # 7)
        elif key == "pl":
            df = _apply_label_map(df, PL_LABEL_MAP)   # 6)
        df_d = df.reset_index() if df.index.name == "科目" else df.copy()
        year_cols = [c for c in df_d.columns if c.startswith("Year")]
        for col in year_cols:
            df_d[col] = df_d[col].apply(_fmt)
        if "科目" in df_d.columns:
            df_d = df_d.set_index("科目")
        out[key] = df_d
    return out


def _render_fs(df: pd.DataFrame):
    """5) 数値列を右寄せで表示"""
    year_cols = [c for c in df.columns if c.startswith("Year")]
    if year_cols:
        st.dataframe(
            df.style.set_properties(subset=year_cols, **{"text-align": "right"}),
            use_container_width=True,
        )
    else:
        st.dataframe(df, use_container_width=True)


def render_pl(d): st.markdown("### 📊 損益計算書（PL）");    _render_fs(d["pl"])
def render_bs(d): st.markdown("### 🏦 貸借対照表（BS）");    _render_fs(d["bs"])
def render_cf(d): st.markdown("### 💸 資金収支計算書（CF）"); _render_fs(d["cf"])


# ============================================================
# 経済探偵レポート 計算ロジック（値のみ返す）
#
# 【修正方針】
#   - CF集計を「売上高・販管費・税」の3科目限定から
#     「預金勘定の純増減（借方増加 - 貸方減少）」ベースに変更
#     → 売却代金・借入・返済・固定資産税など全ての現金移動を捕捉
#   - ROI分母：元入金=0の場合は総投資額（建物+土地+仲介）を使用
#   - 投資回収判定：初期投出金（元入金 or 総投資額）を回収できた月
# ============================================================
def calc_detective_metrics(fs_data: dict, params: SimulationParams, ledger_df: pd.DataFrame) -> dict:
    pl = fs_data["pl"]
    bs = fs_data["bs"]
    total_rent = float(pl.loc["売上高"].sum())           if "売上高"           in pl.index else 0.0
    total_mgmt = float(pl.loc["販売費一般管理費"].sum()) if "販売費一般管理費" in pl.index else 0.0
    mgmt_ratio = total_mgmt / total_rent                 if total_rent != 0    else 0.0
    total_tax  = float(pl.loc["所得税（法人税）"].sum()) if "所得税（法人税）" in pl.index else 0.0
    final_cash = float(bs.loc["預金"].iloc[-1])          if "預金"             in bs.index else 0.0

    # 日付・年月カラムの準備
    ldf = ledger_df.copy()
    if "year" not in ldf.columns or "month" not in ldf.columns:
        dc = next((c for c in ["date", "booking_date", "txn_date"] if c in ldf.columns), None)
        if dc:
            ldf[dc] = pd.to_datetime(ldf[dc])
            ldf["year"]  = ldf[dc].dt.year
            ldf["month"] = ldf[dc].dt.month
        else:
            ldf["year"] = ldf["month"] = 1

    # ── CFテーブルの「営業収支」行を直接使用（表示と一致させる）──
    cf_tbl = fs_data.get("cf", pd.DataFrame())
    # 営業収支の年次値を取得
    op_cf_row = None
    for label in ["営業収支"]:
        if label in cf_tbl.index:
            op_cf_row = cf_tbl.loc[label]
            break

    if op_cf_row is not None:
        # 営業収支が初めてプラスになる年を判定
        pos_year = None
        for col in op_cf_row.index:
            if isinstance(op_cf_row[col], (int, float)) and op_cf_row[col] > 0:
                # col は "Year XXXX" 形式
                try:
                    pos_year = int(str(col).replace("Year ", "").strip())
                except Exception:
                    pass
                if pos_year:
                    break
        pos_s = f"{pos_year}年" if pos_year else "黒字転換なし"

        # 投資回収：累積営業収支が inv_base を超えた年
        cumulative = 0.0
        rec_year = None
        for col in op_cf_row.index:
            try:
                v = float(op_cf_row[col])
                cumulative += v
                yr = int(str(col).replace("Year ", "").strip())
                total_inv_check = (params.property_price_building
                                   + params.property_price_land
                                   + params.brokerage_fee_amount_incl)
                inv_b = params.initial_equity if params.initial_equity > 0 else total_inv_check
                if cumulative >= inv_b:
                    rec_year = yr
                    break
            except Exception:
                continue
        rec_s = f"{rec_year}年" if rec_year else "未回収"

        # NPV は CFテーブルの年次営業収支から
        dr = params.cf_discount_rate or 0.03
        op_vals = []
        for col in op_cf_row.index:
            try:
                op_vals.append(float(op_cf_row[col]))
            except Exception:
                pass
    else:
        pos_s = "データなし"
        rec_s = "データなし"
        op_vals = []
        dr = params.cf_discount_rate or 0.03

    # ROI分母：元入金=0なら総投資額を使用
    total_inv = (params.property_price_building
                 + params.property_price_land
                 + params.brokerage_fee_amount_incl)
    inv_base = params.initial_equity if params.initial_equity > 0 else total_inv

    # ROI
    tp      = final_cash - inv_base
    roi     = tp / inv_base              if inv_base > 0             else 0.0
    ann_roi = roi / params.holding_years if params.holding_years > 0 else 0.0

    # NPV（年次営業収支ベース）
    npv = (
        sum(cf / ((1 + dr) ** (i + 1)) for i, cf in enumerate(op_vals))
        - inv_base
    )

    # 保有期間中の月次現金収支合計（売却含む全キャッシュ）
    op_tot = float(sum(op_vals))

    return {
        "受け取った家賃収入の総額":     total_rent,
        "支払った管理費の総額":         total_mgmt,
        "管理費 ÷ 収入":               mgmt_ratio,
        "支払った税金の総額":           total_tax,
        "営業収支がプラスになる時期":   pos_s,
        "投資回収完了月":               rec_s,
        "売却時に手元に残った金額":     final_cash,
        "全体の投資利回り":             roi,
        "全体の投資利回り年率":         ann_roi,
        "DCF法による現在価値":         npv,
        "借入返済期間中の営業収支合計": op_tot,
    }


def economic_detective_report(fs_data: dict, params: SimulationParams, ledger_df: pd.DataFrame):
    st.subheader("🕵️‍♂️ 経済探偵の分析レポート")
    metrics = calc_detective_metrics(fs_data, params, ledger_df)

    def card(label, value):
        return (f'<div class="bkw-card"><div class="bkw-label">{label}</div>'
                f'<div class="bkw-value">{value}</div></div>')

    def fv(key, val):
        if isinstance(val, str):
            return val
        if "利回り" in key or "÷" in key:
            return f"{val:.1%}"
        return f"{int(val):,} 円"

    # 2) 指定の表示順
    order = [
        "受け取った家賃収入の総額",
        "管理費 ÷ 収入",
        "営業収支がプラスになる時期",
        "売却時に手元に残った金額",
        "借入返済期間中の営業収支合計",
        "支払った管理費の総額",
        "支払った税金の総額",
        "投資回収完了月",
        "全体の投資利回り",
        "全体の投資利回り年率",
        "DCF法による現在価値",
    ]
    cl, cr = st.columns(2)
    for i, k in enumerate(order):
        (cl if i % 2 == 0 else cr).markdown(
            card(k, fv(k, metrics.get(k, ""))), unsafe_allow_html=True
        )


# ============================================================
# シナリオ CSV 生成（入力値）
# ============================================================
def build_scenario_csv(params: SimulationParams, scenario_name: str) -> bytes:
    rows = [
        ("シナリオ名",               scenario_name),
        ("シミュレーション開始日",   str(params.start_date)),
        ("建物価格（税込）",         f"{params.property_price_building:,.0f}"),
        ("土地価格",                 f"{params.property_price_land:,.0f}"),
        ("仲介手数料（税込）",       f"{params.brokerage_fee_amount_incl:,.0f}"),
        ("建物耐用年数",             params.building_useful_life),
        ("初期借入金額",             f"{params.initial_loan.amount:,.0f}" if params.initial_loan else "0"),
        ("借入金利（%）",            f"{params.initial_loan.interest_rate*100:.2f}%" if params.initial_loan else "0%"),
        ("返済期間（年）",           params.initial_loan.years if params.initial_loan else 0),
        ("返済方式",                 "元利均等" if (params.initial_loan and params.initial_loan.repayment_method == "annuity") else "元金均等"),
        ("元入金（自動計算）",       f"{params.initial_equity:,.0f}"),
        ("年間家賃収入（税込）",     f"{params.annual_rent_income_incl:,.0f}"),
        ("非課税割合（%）",          f"{params.non_taxable_proportion*100:.1f}%"),
        ("年間管理費（税込）",       f"{params.annual_management_fee_initial:,.0f}"),
        ("年間修繕費（税込）",       f"{params.repair_cost_annual:,.0f}"),
        ("年間保険料",               f"{params.insurance_cost_annual:,.0f}"),
        ("固定資産税（土地）",       f"{params.fixed_asset_tax_land:,.0f}"),
        ("固定資産税（建物）",       f"{params.fixed_asset_tax_building:,.0f}"),
        ("その他販管費（税込）",     f"{params.other_management_fee_annual:,.0f}"),
        ("消費税率（%）",            f"{params.consumption_tax_rate*100:.1f}%"),
        ("課税主体",                 "個人" if params.entity_type == "individual" else "法人"),
        ("所得税率（%）",            f"{params.income_tax_rate*100:.1f}%"),
        ("法人税率（%）",            f"{params.corporate_tax_rate*100:.1f}%"),
        ("当座借越金利（%）",        f"{params.overdraft_interest_rate*100:.2f}%"),
        ("売却予定年",               params.holding_years),
        ("土地売却額",               f"{params.exit_params.land_exit_price:,.0f}"),
        ("建物売却額（税込）",       f"{params.exit_params.building_exit_price:,.0f}"),
        ("売却費用（税込）",         f"{params.exit_params.exit_cost:,.0f}"),
        ("追加投資件数",             len(params.additional_investments)),
    ]
    for i, inv in enumerate(params.additional_investments, 1):
        rows += [
            (f"追加投資{i}_投資年",        inv.year),
            (f"追加投資{i}_金額（税込）",  f"{inv.amount:,.0f}"),
            (f"追加投資{i}_耐用年数",      inv.life),
            (f"追加投資{i}_借入金額",      f"{inv.loan_amount:,.0f}"),
            (f"追加投資{i}_借入利率（%）", f"{inv.loan_interest_rate*100:.2f}%"),
            (f"追加投資{i}_借入期間",      inv.loan_years),
        ]
    return (pd.DataFrame(rows, columns=["項目", "入力値"])
            .to_csv(index=False, encoding="utf-8-sig")
            .encode("utf-8-sig"))


# ============================================================
# 結果 Excel 生成
# 1) 入力条件シートに経済探偵レポートを追記
# 2) 指定項目・順序
# 5) 数値列を右寄せ（openpyxlスタイル適用）
# ============================================================
def build_result_excel(
    fs_data: dict,
    ledger_df: pd.DataFrame,
    params: SimulationParams,
    scenario_name: str,
    metrics: dict,
) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet1: 入力条件 + 経済探偵レポート ──────────────
        input_rows = [
            ("■ 入力条件", ""),
            ("シナリオ名",               scenario_name),
            ("シミュレーション開始日",   str(params.start_date)),
            ("建物価格（税込）",         params.property_price_building),
            ("土地価格",                 params.property_price_land),
            ("仲介手数料（税込）",       params.brokerage_fee_amount_incl),
            ("建物耐用年数（年）",       params.building_useful_life),
            ("初期借入金額",             params.initial_loan.amount if params.initial_loan else 0),
            ("借入金利（%）",            params.initial_loan.interest_rate * 100 if params.initial_loan else 0),
            ("返済期間（年）",           params.initial_loan.years if params.initial_loan else 0),
            ("返済方式",                 "元利均等" if (params.initial_loan and params.initial_loan.repayment_method == "annuity") else "元金均等"),
            ("元入金",                   params.initial_equity),
            ("年間家賃収入（税込）",     params.annual_rent_income_incl),
            ("非課税割合（%）",          params.non_taxable_proportion * 100),
            ("年間管理費（税込）",       params.annual_management_fee_initial),
            ("年間修繕費（税込）",       params.repair_cost_annual),
            ("年間保険料",               params.insurance_cost_annual),
            ("固定資産税（土地）",       params.fixed_asset_tax_land),
            ("固定資産税（建物）",       params.fixed_asset_tax_building),
            ("その他販管費（税込）",     params.other_management_fee_annual),
            ("消費税率（%）",            params.consumption_tax_rate * 100),
            ("課税主体",                 "個人" if params.entity_type == "individual" else "法人"),
            ("所得税率（%）",            params.income_tax_rate * 100),
            ("法人税率（%）",            params.corporate_tax_rate * 100),
            ("当座借越金利（%）",        params.overdraft_interest_rate * 100),
            ("売却予定年",               params.holding_years),
            ("土地売却額",               params.exit_params.land_exit_price),
            ("建物売却額（税込）",       params.exit_params.building_exit_price),
            ("売却費用（税込）",         params.exit_params.exit_cost),
            ("追加投資件数",             len(params.additional_investments)),
        ]
        for i, inv in enumerate(params.additional_investments, 1):
            input_rows += [
                (f"追加投資{i}_投資年",        inv.year),
                (f"追加投資{i}_金額（税込）",  inv.amount),
                (f"追加投資{i}_耐用年数",      inv.life),
                (f"追加投資{i}_借入金額",      inv.loan_amount),
                (f"追加投資{i}_借入利率（%）", inv.loan_interest_rate * 100),
                (f"追加投資{i}_借入期間",      inv.loan_years),
            ]

        # 経済探偵レポートをこのシートに追記
        input_rows += [("", ""), ("■ 経済探偵レポート", "")]
        det_order = [
            "受け取った家賃収入の総額",
            "管理費 ÷ 収入",
            "営業収支がプラスになる時期",
            "売却時に手元に残った金額",
            "借入返済期間中の営業収支合計",
            "支払った管理費の総額",
            "支払った税金の総額",
            "投資回収完了月",
            "全体の投資利回り",
            "全体の投資利回り年率",
            "DCF法による現在価値",
        ]
        for k in det_order:
            v = metrics.get(k, "")
            if isinstance(v, float):
                v = f"{v:.1%}" if ("利回り" in k or "÷" in k) else round(v)
            input_rows.append((k, v))

        pd.DataFrame(input_rows, columns=["項目", "値"]).to_excel(
            writer, sheet_name="入力条件・分析サマリー", index=False
        )

        # ── Sheet2〜4: PL / BS / CF（数値のまま出力・右寄せ後処理）──
        for key, sname, lmap in [
            ("pl", "損益計算書PL",  PL_LABEL_MAP),
            ("bs", "貸借対照表BS",  {}),
            ("cf", "資金収支CF",    CF_LABEL_MAP),
        ]:
            if key not in fs_data:
                continue
            df = _apply_label_map(fs_data[key].copy(), lmap)
            df.index.name = "科目"
            df.reset_index().to_excel(writer, sheet_name=sname, index=False)

        # ── Sheet5: 全仕訳 ──────────────────────────────────
        ledger_df.to_excel(writer, sheet_name="全仕訳", index=False)

        # ── 数値列の右寄せスタイル適用 ──────────────────────
        wb = writer.book
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(cell.value, float) and cell.value != int(cell.value):
                            pass  # 割合は文字列で出力済み
                        else:
                            cell.number_format = "#,##0"

    return buf.getvalue()


# ============================================================
# 追加投資入力（expander）
# ============================================================
def _setup_additional_investments_internal(
    num_investments: int,
    exit_year: int,
) -> List[AdditionalInvestmentParams]:
    investments: List[AdditionalInvestmentParams] = []
    if num_investments == 0:
        return investments
    for i in range(1, num_investments + 1):
        with st.sidebar.expander(f"第{i}回 追加投資", expanded=False):
            inv_year = st.number_input(
                "投資年", min_value=1, max_value=exit_year, value=1, step=1,
                key=f"aiy_{i}",
            )
            inv_amount = st.number_input(
                "投資金額（税込）", min_value=0.0, max_value=999_999_999.0,
                step=100_000.0, format="%.0f", key=f"aia_{i}",
            )
            inv_life = st.number_input(
                "耐用年数", min_value=1, max_value=100, value=15, step=1, key=f"ail_{i}",
            )
            ld = (inv_amount == 0)
            inv_loan = st.number_input(
                "付随借入金額",
                min_value=0.0,
                max_value=float(inv_amount) if inv_amount > 0 else 0.0,
                step=100_000.0, format="%.0f", key=f"ailn_{i}", disabled=ld,
            )
            dd = ld or (inv_loan == 0)
            inv_loan_years = st.number_input(
                "借入期間（年）", min_value=0, max_value=50, value=0, step=1,
                key=f"aily_{i}", disabled=dd,
            )
            inv_loan_rate = st.number_input(
                "借入利率（%）", min_value=0.0, max_value=100.0, value=2.5, step=0.01,
                key=f"ailr_{i}", disabled=dd,
            ) / 100
            if inv_amount > 0:
                investments.append(
                    AdditionalInvestmentParams(
                        year=int(inv_year), amount=float(inv_amount), life=int(inv_life),
                        loan_amount=float(inv_loan), loan_years=int(inv_loan_years),
                        loan_interest_rate=float(inv_loan_rate),
                    )
                )
    return investments


# ============================================================
# サイドバー全体（expander方式）
# ============================================================
def setup_sidebar() -> SimulationParams:
    C = "%.0f"
    st.sidebar.markdown("## 🛠 入力欄")
    st.sidebar.caption("各セクションをクリックして展開してください。")

    # ── 1. 物件情報 ──────────────────────────────────────────
    with st.sidebar.expander("🏠 1. 物件情報", expanded=True):
        start_date = st.date_input(
            "シミュレーション開始日", value=datetime.date(2026, 1, 1), key="ssd",
        )
        price_bld = st.number_input(
            "建物価格（税込）", min_value=0.0, max_value=999_999_999.0,
            value=50_000_000.0, step=100_000.0, format=C,
        )
        price_land = st.number_input(
            "土地価格（非課税）", min_value=0.0, max_value=999_999_999.0,
            value=30_000_000.0, step=100_000.0, format=C,
        )
        brokerage = st.number_input(
            "仲介手数料（税込）", min_value=0.0, max_value=999_999_999.0,
            value=3_300_000.0, step=10_000.0, format=C,
        )
        bld_zero  = (price_bld  == 0.0)
        land_zero = (price_land == 0.0)
        bul = st.number_input(
            "建物耐用年数（年）", min_value=1, max_value=55, value=47, step=1,
            key="bul", disabled=bld_zero,
        )
        bag = st.number_input(
            "建物築年数（年）※参考", min_value=0, max_value=60, value=5, step=1, key="bag",
        )

    # ── 2. 融資条件 ──────────────────────────────────────────
    total_inv = price_bld + price_land + brokerage
    with st.sidebar.expander("💰 2. 融資条件", expanded=True):
        loan_amt = st.number_input(
            "初期借入金額",
            min_value=0.0,
            max_value=float(total_inv) if total_inv > 0 else 999_999_999.0,
            value=min(70_000_000.0, float(total_inv)),
            step=100_000.0, format=C,
        )
        loan_dis = (loan_amt == 0)
        loan_yrs = st.number_input(
            "返済期間（年）", min_value=1, max_value=50, value=30, step=1, disabled=loan_dis,
        )
        loan_rate = st.number_input(
            "借入金利（年率 %）", min_value=0.0, max_value=100.0, value=2.5, step=0.01,
            disabled=loan_dis,
        ) / 100
        repay_method = st.selectbox(
            "返済方式",
            options=["annuity", "equal_principal"],
            format_func=lambda x: "元利均等" if x == "annuity" else "元金均等",
            key="rm", disabled=loan_dis,
        )
        initial_loan = (
            LoanParams(amount=loan_amt, interest_rate=loan_rate,
                       years=int(loan_yrs), repayment_method=repay_method)
            if loan_amt > 0 else None
        )
        equity = float(max(total_inv - loan_amt, 0.0))
        st.metric("元入金（自動計算）", f"{equity:,.0f} 円")

    # ── 3. 収益・費用 ─────────────────────────────────────────
    with st.sidebar.expander("🏢 3. 収益・費用", expanded=False):
        annual_rent = st.number_input(
            "年間家賃収入（税込）", min_value=0.0, max_value=999_999_999.0,
            value=3_600_000.0, step=10_000.0, format=C,
        )
        ntx_pct = st.number_input(
            "非課税割合（%）※住宅割合", min_value=0.0, max_value=100.0, value=0.0, step=5.0,
        )
        non_tax_prop = ntx_pct / 100.0
        mgmt_fee = st.number_input(
            "年間管理費（税込）", min_value=0.0, max_value=999_999_999.0,
            value=1_200_000.0, step=10_000.0, format=C,
        )
        repair_cost = st.number_input(
            "年間修繕費（税込）", min_value=0.0, max_value=999_999_999.0,
            value=300_000.0, step=10_000.0, format=C,
        )
        insurance = st.number_input(
            "年間保険料（非課税）", min_value=0.0, max_value=999_999_999.0,
            value=100_000.0, step=10_000.0, format=C,
        )
        fa_tax_land = st.number_input(
            "固定資産税（土地）", min_value=0.0, max_value=999_999_999.0,
            value=150_000.0, step=10_000.0, format=C, disabled=land_zero,
        )
        fa_tax_bld = st.number_input(
            "固定資産税（建物）", min_value=0.0, max_value=999_999_999.0,
            value=150_000.0, step=10_000.0, format=C, disabled=bld_zero,
        )
        other_fee = st.number_input(
            "その他販管費（税込・年額）", min_value=0.0, max_value=999_999_999.0,
            value=0.0, step=10_000.0, format=C,
        )

    # ── 4. 税率設定 ───────────────────────────────────────────
    with st.sidebar.expander("📊 4. 税率設定", expanded=False):
        vat_rate = st.number_input(
            "消費税率（%）", min_value=0.0, max_value=100.0, value=10.0, step=0.1,
        ) / 100
        entity_type = st.selectbox(
            "課税主体",
            options=["individual", "corporate"],
            format_func=lambda x: "個人" if x == "individual" else "法人",
            key="et",
        )
        tax_rate = st.number_input(
            "所得税率（%）" if entity_type == "individual" else "法人税率（%）",
            min_value=0.0, max_value=100.0, value=30.0, step=0.1,
        ) / 100
        od_rate = st.number_input(
            "当座借越金利（%）", min_value=0.0, max_value=100.0, value=5.0, step=0.01,
        ) / 100

    # ── 5. 出口設定 ───────────────────────────────────────────
    with st.sidebar.expander("📉 5. 出口設定", expanded=True):
        exit_year = st.number_input("売却予定年", min_value=1, max_value=100, value=5, step=1)
        hy = int(exit_year)
        c1, c2 = st.columns(2)
        with c1:
            lep = st.number_input(
                "土地売却額（非課税）",
                min_value=0.0, max_value=999_999_999.0,
                value=float(price_land), step=100_000.0, format="%.0f",
                key="lep", disabled=land_zero,
            )
        with c2:
            bep = st.number_input(
                "建物売却額（税込）",
                min_value=0.0, max_value=999_999_999.0,
                value=float(price_bld), step=100_000.0, format="%.0f",
                key="bep", disabled=bld_zero,
            )
        ec = st.number_input(
            "売却費用（税込）",
            min_value=0.0, max_value=999_999_999.0,
            value=0.0, step=10_000.0, format="%.0f", key="ec",
        )

    exit_params = ExitParams(
        exit_year=hy,
        land_exit_price=float(lep) if not land_zero else 0.0,
        building_exit_price=float(bep) if not bld_zero else 0.0,
        exit_cost=float(ec),
    )

    # ── 6. 追加投資 ───────────────────────────────────────────
    with st.sidebar.expander("➕ 6. 追加投資", expanded=False):
        num_inv = st.number_input(
            "追加投資回数（最大5）", min_value=0, max_value=5, value=0, step=1,
        )
    additional_investments = _setup_additional_investments_internal(int(num_inv), hy)

    return SimulationParams(
        property_price_building=float(price_bld),
        property_price_land=float(price_land),
        brokerage_fee_amount_incl=float(brokerage),
        building_useful_life=int(bul),
        building_age=int(bag),
        holding_years=hy,
        initial_loan=initial_loan,
        initial_equity=equity,
        rent_setting_mode="AMOUNT",
        target_cap_rate=0.0,
        annual_rent_income_incl=float(annual_rent),
        annual_management_fee_initial=float(mgmt_fee),
        repair_cost_annual=float(repair_cost),
        insurance_cost_annual=float(insurance),
        fixed_asset_tax_land=float(fa_tax_land) if not land_zero else 0.0,
        fixed_asset_tax_building=float(fa_tax_bld) if not bld_zero else 0.0,
        other_management_fee_annual=float(other_fee),
        management_fee_rate=0.0,
        consumption_tax_rate=float(vat_rate),
        non_taxable_proportion=float(non_tax_prop),
        overdraft_interest_rate=float(od_rate),
        cf_discount_rate=0.0,
        exit_params=exit_params,
        additional_investments=additional_investments,
        start_date=start_date,
        entity_type=entity_type,
        income_tax_rate=tax_rate if entity_type == "individual" else 0.0,
        corporate_tax_rate=tax_rate if entity_type == "corporate"  else 0.0,
    )


# ============================================================
# 入力バリデーション
# ============================================================
def validate_params(params: SimulationParams) -> list:
    errors = []
    if params.property_price_building == 0 and params.property_price_land == 0:
        errors.append("建物価格と土地価格の両方が 0 です。どちらか一方を入力してください。")
    total = (params.property_price_building
             + params.property_price_land
             + params.brokerage_fee_amount_incl)
    if params.initial_loan and params.initial_loan.amount > total:
        errors.append(
            f"借入金額（{params.initial_loan.amount:,.0f}）が"
            f"物件価格合計（{total:,.0f}）を超えています。"
        )
    for i, inv in enumerate(params.additional_investments, 1):
        if inv.year > params.holding_years:
            errors.append(
                f"追加投資 第{i}回の投資年（{inv.year}年）が"
                f"売却予定年（{params.holding_years}年）を超えています。"
            )
        if inv.loan_amount > inv.amount:
            errors.append(
                f"追加投資 第{i}回の借入金額が投資金額を超えています。"
            )
    return errors


# ============================================================
# メイン
# ============================================================
def main():
    st.set_page_config(layout="wide", page_title="BKW Invest Sim (Amelia v4)")
    inject_global_css()
    st.title("💰 BKW 不動産投資シミュレーション")
    st.caption("仕様書 v4訂正済準拠版")

    params = setup_sidebar()

    # ── 4) 前提条件サマリー（3列、全18項目）────────────────────
    st.markdown(
        '<div class="bkw-section-title">📋 シミュレーション前提条件（入力値確認）</div>',
        unsafe_allow_html=True,
    )

    def card(label, value):
        return (f'<div class="bkw-card"><div class="bkw-label">{label}</div>'
                f'<div class="bkw-value">{value}</div></div>')

    add_inv_total = sum(inv.amount for inv in params.additional_investments)

    summary_items = [
        ("建物価格（税込）",     f"{params.property_price_building:,.0f}"),
        ("土地価格",             f"{params.property_price_land:,.0f}"),
        ("仲介手数料（税込）",   f"{params.brokerage_fee_amount_incl:,.0f}"),
        ("初期借入金額",         f"{params.initial_loan.amount:,.0f}" if params.initial_loan else "なし"),
        ("借入金利（%）",        f"{params.initial_loan.interest_rate*100:.2f}%" if params.initial_loan else "—"),
        ("返済期間（年）",       f"{params.initial_loan.years} 年" if params.initial_loan else "—"),
        ("非課税割合（%）",      f"{params.non_taxable_proportion*100:.0f}%"),
        ("年間管理費（税込）",   f"{params.annual_management_fee_initial:,.0f}"),
        ("年間修繕費（税込）",   f"{params.repair_cost_annual:,.0f}"),
        ("年間保険料",           f"{params.insurance_cost_annual:,.0f}"),
        ("固定資産税（土地）",   f"{params.fixed_asset_tax_land:,.0f}"),
        ("固定資産税（建物）",   f"{params.fixed_asset_tax_building:,.0f}"),
        ("その他販管費（税込）", f"{params.other_management_fee_annual:,.0f}"),
        ("売却予定年",           f"{params.holding_years} 年"),
        ("土地売却額",           f"{params.exit_params.land_exit_price:,.0f}"),
        ("建物売却額（税込）",   f"{params.exit_params.building_exit_price:,.0f}"),
        ("売却費用（税込）",     f"{params.exit_params.exit_cost:,.0f}"),
        ("追加設備投資（税込）", f"{add_inv_total:,.0f}"),
    ]
    cols = st.columns(3)
    for i, (label, value) in enumerate(summary_items):
        cols[i % 3].markdown(card(label, value), unsafe_allow_html=True)

    if params.additional_investments:
        st.markdown(
            '<div class="bkw-section-title">➕ 追加投資の詳細</div>',
            unsafe_allow_html=True,
        )
        inv_cols = st.columns(5)
        for idx, inv in enumerate(params.additional_investments):
            with inv_cols[idx % 5]:
                st.markdown(
                    f'<div class="bkw-card" style="min-height:190px;padding:10px;">'
                    f'<div class="bkw-label">第{idx+1}回 追加投資</div>'
                    f'<div class="bkw-value" style="font-size:0.95rem;text-align:left;">'
                    f'投資年：{inv.year} 年目<br>'
                    f'投資金額：{inv.amount:,.0f} 円<br>'
                    f'耐用年数：{inv.life} 年<br>'
                    f'借入金額：{inv.loan_amount:,.0f} 円<br>'
                    f'借入利率：{inv.loan_interest_rate:.2%}<br>'
                    f'返済年数：{inv.loan_years} 年'
                    f'</div></div>',
                    unsafe_allow_html=True,
                )

    # ── シナリオ名 ────────────────────────────────────────────
    st.markdown("---")
    scenario_name = st.text_input(
        "📝 シナリオ名（最大20文字）", max_chars=20, placeholder="例：都心RC 5年保有",
    )

    # ── 免責同意 ──────────────────────────────────────────────
    st.markdown(
        "**免責事項**\n\n"
        "本ツールは参考計算を目的としたシミュレーションです。"
        "実際の税務・投資結果を保証するものではありません。"
    )
    disclaimer_ok = st.checkbox("内容を理解した上で利用します", value=False, key="disclaimer")

    # ── バリデーション ────────────────────────────────────────
    errors = validate_params(params)
    for e in errors:
        st.error(f"⚠️ {e}")

    # ── 実行ボタン ────────────────────────────────────────────
    run_disabled = (not disclaimer_ok) or (len(errors) > 0)
    run_clicked  = st.button(
        "▶︎ シミュレーション実行",
        type="primary", use_container_width=True, disabled=run_disabled,
    )
    if not disclaimer_ok:
        st.caption("※ 免責事項にチェックを入れると実行できます。")
    if errors:
        st.caption("※ 入力エラーを修正すると実行できます。")

    # ── 実行処理 ──────────────────────────────────────────────
    if run_clicked:
        try:
            with st.spinner("計算中..."):
                sim = Simulation(params, params.start_date)
                sim.run()

            ledger_df        = sim.ledger.get_df()
            ledger_df_sorted = ledger_df.sort_values(["date", "id"]).reset_index(drop=True)

            fs_builder = FinancialStatementBuilder(sim.ledger)
            fs_data    = fs_builder.build()
            display_fs = create_display_dataframes(fs_data)

            diff = fs_data.get("balance_diff", 0)
            if fs_data.get("is_balanced", False):
                st.success(f"✅ 貸借合致（差額 {diff:.0f} 円）")
            else:
                st.error(f"❌ 貸借不一致（差額 {diff:.0f} 円）")

            # 経済探偵メトリクス（ダウンロードとUIで共用）
            metrics = calc_detective_metrics(fs_data, params, ledger_df_sorted)

            # ── ダウンロードボタン ────────────────────────────
            now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    "📥 入力条件CSV をダウンロード",
                    data=build_scenario_csv(params, scenario_name),
                    file_name=f"bkw_sim_input_{now_str}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with dl2:
                st.download_button(
                    "📊 演算結果Excel をダウンロード",
                    data=build_result_excel(fs_data, ledger_df_sorted, params, scenario_name, metrics),
                    file_name=f"bkw_sim_result_{now_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            # 経済探偵レポート（UI表示）
            economic_detective_report(fs_data, params, ledger_df_sorted)

            st.session_state["display_fs"]       = display_fs
            st.session_state["ledger_df_sorted"] = ledger_df_sorted

        except Exception as e:
            st.error(f"シミュレーションエラー: {str(e)}")
            st.code(traceback.format_exc())
            return

    # ── 財務三表タブ ──────────────────────────────────────────
    if "display_fs" in st.session_state:
        dfs  = st.session_state["display_fs"]
        ldfs = st.session_state["ledger_df_sorted"]
        tabs = st.tabs([
            "📊 損益計算書（PL）",
            "🏦 貸借対照表（BS）",
            "💸 資金収支（CF）",
            "📒 全仕訳",
        ])
        with tabs[0]: render_pl(dfs)
        with tabs[1]: render_bs(dfs)
        with tabs[2]: render_cf(dfs)
        with tabs[3]:
            amt_cols = ["amount"] if "amount" in ldfs.columns else []
            st.dataframe(
                ldfs.style.set_properties(subset=amt_cols, **{"text-align": "right"})
                if amt_cols else ldfs,
                use_container_width=True,
            )


if __name__ == "__main__":
    main()